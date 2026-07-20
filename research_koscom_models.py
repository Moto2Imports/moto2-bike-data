#!/usr/bin/env python3
"""
research_koscom_models.py — STANDALONE research tool (NOT part of the daily
scraper pipeline). Catalogues the real model names koscom/BDS uses per make, so
Tim's BDS-named spreadsheet can be corrected against koscom's actual labelling
(RGV250, NSR50, NSR80 already known to differ; this finds the rest).

For each make it:
  1. Browses koscom via ?manuf=<make>&force=1 (reusing the browse pagination
     helpers in koscom_scraper_v3 — _paginate_search). No max_year/model filter:
     we want every model name, all years.
  2. Fetches each listing and keeps BDS auctions only (extract_auction_house,
     the same BDS-only rule the scraper uses).
  3. Reads the model name as koscom labels it (extract_model_name).
  4. Dedupes to distinct model names per make (with a listing count each).
  5. Writes a clean .xlsx: a "Models" sheet (Make / Model / Listings seen) plus
     a "Summary" sheet (per-make totals).

Because the model name lives on the DETAIL page, discovering a make's full model
set means fetching listings until new names stop appearing. `--stale-limit`
stops a make after N consecutive BDS listings yield no new model (convergence);
set 0 to scan exhaustively. Accuracy-first: the default errs large.

⚠️ Reuses extract_model_name / extract_auction_house, whose selectors are
UNVERIFIED against live koscom markup. "(model not found)" rows flag listings
whose model didn't parse — if there are many, the selector needs adjusting from
a sample of real listing HTML. This is a research aid, not a customer-facing
feed, so that risk is acceptable here.

    python3 research_koscom_models.py                       # all 7 makes
    python3 research_koscom_models.py --makes Honda Bimota  # subset
    python3 research_koscom_models.py --stale-limit 0       # exhaustive
    python3 research_koscom_models.py -o koscom_models.xlsx # output path
"""
import argparse
import time
from collections import OrderedDict

from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from koscom_scraper_v3 import (
    MAX_PAGES_PER_MAKE,
    REQUEST_DELAY_SECONDS,
    SEARCH_TIMEOUT,
    KoscomScraperV3,
    extract_auction_house,
    extract_model_name,
)

MAKES = ["Honda", "Suzuki", "Kawasaki", "Yamaha", "Bimota", "BMW", "Aprilia"]
MODEL_NOT_FOUND = "(model not found)"


def collect_models_for_make(scraper, make, stale_limit=120, max_per_make=0,
                            fetch_delay=REQUEST_DELAY_SECONDS, log=print):
    """Return {'models': OrderedDict(name->count), 'scanned', 'non_bds',
    'stopped_early'} for one make. `models` is in first-seen order."""
    models = OrderedDict()
    scanned = non_bds = stale = 0
    stopped_early = False
    for _lid, url, _cc in scraper._paginate_search(
            {"manuf": make, "force": 1}, exclude_seen=False,
            max_pages=MAX_PAGES_PER_MAKE):
        try:
            resp = scraper.session.get(url, timeout=SEARCH_TIMEOUT)
            resp.encoding = "utf-8"
        except Exception as e:
            log(f"[ERROR] {make} {url}: {e}")
            continue
        soup = BeautifulSoup(resp.text, "html.parser")
        page_text = soup.get_text(" ", strip=True)
        if extract_auction_house(page_text) is None:      # BDS-only, reused rule
            non_bds += 1
            continue
        scanned += 1
        model = extract_model_name(soup, make) or MODEL_NOT_FOUND
        if model in models:
            models[model] += 1
            stale += 1
        else:
            models[model] = 1
            stale = 0
            log(f"[{make}] + {model!r}  ({len(models)} distinct, {scanned} scanned)")
        if max_per_make and scanned >= max_per_make:
            stopped_early = True
            log(f"[{make}] hit --max-per-make {max_per_make}")
            break
        if stale_limit and stale >= stale_limit:
            stopped_early = True
            log(f"[{make}] converged — {stale} BDS listings with no new model")
            break
        if fetch_delay:
            time.sleep(fetch_delay)
    log(f"[{make}] done: {len(models)} distinct model(s), {scanned} BDS listings, "
        f"{non_bds} non-BDS skipped{' (stopped early)' if stopped_early else ''}")
    return {"models": models, "scanned": scanned, "non_bds": non_bds,
            "stopped_early": stopped_early}


# ---------------------------------------------------------------- workbook out
_ARIAL = "Arial"
_HEADER_FONT = Font(name=_ARIAL, bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill("solid", fgColor="404040")
_BODY_FONT = Font(name=_ARIAL)


def _style_header(ws, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(ncols)}1"


def build_workbook(results, makes, path):
    """results: {make: collect_models_for_make(...) dict}. Writes `path`."""
    wb = openpyxl.Workbook()

    # -- Models sheet: one row per distinct (make, model), Make/Model columns.
    ws = wb.active
    ws.title = "Models"
    ws.append(["Make", "Model (koscom naming)", "Listings seen"])
    for make in makes:
        data = results.get(make, {"models": {}})
        # sort models alphabetically, but keep the not-found sentinel last.
        for model in sorted(data["models"],
                            key=lambda m: (m == MODEL_NOT_FOUND, m.lower())):
            ws.append([make, model, data["models"][model]])
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = _BODY_FONT
    _style_header(ws, 3)
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 14

    # -- Summary sheet: per-make totals.
    sm = wb.create_sheet("Summary")
    sm.append(["Make", "Distinct models", "BDS listings scanned",
               "Non-BDS skipped", "Coverage"])
    for make in makes:
        d = results.get(make)
        if d is None:
            sm.append([make, 0, 0, 0, "not run"])
            continue
        coverage = "stopped early" if d["stopped_early"] else "exhaustive"
        sm.append([make, len(d["models"]), d["scanned"], d["non_bds"], coverage])
    for row in sm.iter_rows(min_row=2):
        for cell in row:
            cell.font = _BODY_FONT
    _style_header(sm, 5)
    for col, w in zip("ABCDE", (12, 16, 22, 16, 16)):
        sm.column_dimensions[col].width = w

    wb.save(path)
    return path


def main():
    ap = argparse.ArgumentParser(description="Catalogue koscom's real per-make model names (research).")
    ap.add_argument("--makes", nargs="+", default=MAKES, help="makes to scan")
    ap.add_argument("-o", "--out", default="koscom_models_research.xlsx", help="output .xlsx path")
    ap.add_argument("--stale-limit", type=int, default=120,
                    help="stop a make after N BDS listings with no new model (0 = exhaustive)")
    ap.add_argument("--max-per-make", type=int, default=0, help="cap BDS listings scanned per make (0 = no cap)")
    args = ap.parse_args()

    scraper = KoscomScraperV3()
    results = {}
    for make in args.makes:
        results[make] = collect_models_for_make(
            scraper, make, stale_limit=args.stale_limit, max_per_make=args.max_per_make)
    path = build_workbook(results, args.makes, args.out)
    total = sum(len(r["models"]) for r in results.values())
    print(f"\n[SAVED] {path} — {total} distinct model(s) across {len(args.makes)} make(s)")


if __name__ == "__main__":
    main()
