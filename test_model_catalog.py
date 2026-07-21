#!/usr/bin/env python3
"""
Self-tests for koscom_model_catalog.py — no live access. Covers token
extraction (several source shapes + failure), ModelName#Count parsing (wrapped
callback, unicode/Γ/spaces/hyphens, dedupe), the POST request shape (token in
URL, manuf + optional house in the form), and the .xlsx output.

    python3 test_model_catalog.py
"""
import os
import tempfile
import types

import openpyxl

import koscom_model_catalog as mc


# ------------------------------------------------------------ token extraction
def test_token_from_js_assignment():
    assert mc.extract_ajx_token("var ajx = 'a1b2c3d4';") == "a1b2c3d4"
    assert mc.extract_ajx_token('config = {ajx:"Zx99Kp77"};') == "Zx99Kp77"


def test_token_from_url_in_source():
    html = "<script>u='/bike_st?file=ajxModel&ajx=TOK12345&x=1'</script>"
    assert mc.extract_ajx_token(html) == "TOK12345"


def test_token_from_hidden_input_either_order():
    assert mc.extract_ajx_token('<input name="ajx" value="Hidden123">') == "Hidden123"
    assert mc.extract_ajx_token('<input value="Hidden456" name="ajx">') == "Hidden456"


def test_token_absent_returns_none():
    assert mc.extract_ajx_token("<html>no token here</html>") is None


def test_fetch_token_raises_when_missing():
    s = types.SimpleNamespace(get=lambda u, timeout=None:
                              types.SimpleNamespace(text="<html/>", encoding="utf-8"))
    try:
        mc.fetch_ajx_token(s)
        assert False, "expected RuntimeError"
    except RuntimeError:
        pass


def test_fetch_token_reads_fresh_from_stats_page():
    calls = []
    s = types.SimpleNamespace(get=lambda u, timeout=None: (
        calls.append(u),
        types.SimpleNamespace(text="ajx:'Fresh999'", encoding="utf-8"))[1])
    assert mc.fetch_ajx_token(s) == "Fresh999"
    assert calls == [mc.STATS_URL]                # loaded the stats page, not cached


# --------------------------------------------------------------- pair parsing
def test_parse_basic_pairs():
    assert mc.parse_model_counts("CBR400RR#31=VFR400R#14=RVF400#6") == [
        ("CBR400RR", 31, True), ("VFR400R", 14, True), ("RVF400", 6, True)]


def test_parse_wrapped_callback_and_unicode():
    payload = "ajxModelCB('RGV250Γ#4=GSX-R400#11=NSR250 SP#2')"
    assert mc.parse_model_counts(payload) == [
        ("RGV250Γ", 4, True), ("GSX-R400", 11, True), ("NSR250 SP", 2, True)]


def test_parse_dedupes_summing_counts():
    assert mc.parse_model_counts("A#3=B#1=A#2") == [("A", 5, True), ("B", 1, True)]


def test_parse_empty_or_garbage():
    assert mc.parse_model_counts("") == []
    assert mc.parse_model_counts("no pairs at all") == []


def test_parse_strips_no_title_marker_and_flags():
    # The no-title suffix is stripped from the model name and flagged; the same
    # base model with and without the marker becomes two rows.
    payload = "NSR250R-1#20=NSR250R-1 SHO LOUIS NOT EQUIPPED#5"
    assert mc.parse_model_counts(payload) == [
        ("NSR250R-1", 20, True), ("NSR250R-1", 5, False)]


# ---------------------------------------------------------------- request shape
def _capture_post():
    calls = {}

    def post(url, data=None, headers=None, timeout=None):
        calls["url"] = url
        calls["data"] = data
        return types.SimpleNamespace(text="CBR400RR#31=VFR400R#14", encoding="utf-8")

    return types.SimpleNamespace(post=post), calls


def test_request_puts_token_in_url_and_manuf_in_form():
    s, calls = _capture_post()
    pairs = mc.fetch_models_for_make(s, "Honda", "TOK1", house=None)
    assert "ajx=TOK1" in calls["url"] and "file=ajxModel" in calls["url"]
    assert calls["data"] == {"manuf": "Honda"}
    assert pairs == [("CBR400RR", 31, True), ("VFR400R", 14, True)]


def test_house_param_added_when_set():
    s, calls = _capture_post()
    mc.fetch_models_for_make(s, "Honda", "TOK1", house="BDS")
    assert calls["data"] == {"manuf": "Honda", "house": "BDS"}


# ---------------------------------------------------------------- workbook out
def test_build_workbook_structure_and_sorting():
    results = {
        "Honda": [("VFR400R", 14, True), ("CBR400RR", 31, True),
                  ("RVF400", 6, True), ("CBR400RR", 3, False)],   # a no-title CBR bucket
        "Bimota": [],
    }
    with tempfile.TemporaryDirectory() as d:
        path = os.path.join(d, "cat.xlsx")
        mc.build_workbook(results, ["Honda", "Bimota"], path, house="BDS")
        wb = openpyxl.load_workbook(path)
        assert wb.sheetnames == ["Models", "Summary"]
        ws = wb["Models"]
        assert [c.value for c in ws[1]] == [
            "Make", "Model (koscom naming)", "Has title", "Listings"]
        rows = [(r[0].value, r[1].value, r[2].value, r[3].value)
                for r in ws.iter_rows(min_row=2)]
        assert rows == [                                   # sorted by model, title before no-title
            ("Honda", "CBR400RR", "Yes", 31),
            ("Honda", "CBR400RR", "NO", 3),
            ("Honda", "RVF400", "Yes", 6),
            ("Honda", "VFR400R", "Yes", 14),
        ], rows
        assert ws["A2"].font.name == "Arial"
        sm = wb["Summary"]
        by_make = {r[0].value: (r[1].value, r[2].value, r[3].value)
                   for r in sm.iter_rows(min_row=2, max_row=3)}
        assert by_make["Honda"] == (4, 54, 3)              # 4 rows, 54 listings, 3 no-title
        assert by_make["Bimota"] == (0, 0, 0)


if __name__ == "__main__":
    fns = [v for k, v in sorted(globals().items()) if k.startswith("test_")]
    for fn in fns:
        fn()
        print(f"ok  {fn.__name__}")
    print(f"\n{len(fns)} test(s) passed")
