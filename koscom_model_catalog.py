#!/usr/bin/env python3
"""
koscom_model_catalog.py — STANDALONE research tool. Pulls koscom's authoritative
per-make model catalog from its stats endpoint (one request per make), replacing
the old listing-crawl approach (research_koscom_models.py).

Endpoint (discovered): the stats page exposes a per-page-load `ajx` token, and
    POST /bike_st?file=ajxModel&ajx=<token>   form: manuf=<Make>
returns that make's entire distinct model list in one response — a JS callback
carrying 'ModelName#Count=ModelName#Count=...'.

Flow:
  1. GET the /bike_st stats page and extract the CURRENT `ajx` token from its
     source. The token is session/page-load specific and expires — it is ALWAYS
     read fresh here, never hardcoded.
  2. For each make, POST manuf=<Make> with that token.
  3. Parse the ModelName#Count pairs.
  4. Write the same clean .xlsx as before (Models sheet + Summary sheet).

⚠️ UNVERIFIED against live markup (site unreachable from the build env):
  * AJX_TOKEN_PATTERNS — the shapes we look for in the stats-page source. If the
    token sits in a form the patterns miss, --ajx lets you pass one for a debug
    run, and the patterns are easy to extend once you can view-source the page.
  * --house (BDS pre-filter): the request param koscom's BDS auction-house toggle
    adds is NOT confirmed. `--house BDS` sends `house=BDS`; verify by toggling BDS
    in koscom's own UI and checking the request params (DevTools), then adjust
    the value/name if it differs.

    python3 koscom_model_catalog.py                        # all 7 makes, full catalog
    python3 koscom_model_catalog.py --house BDS            # BDS-only (unverified param)
    python3 koscom_model_catalog.py --makes Honda Bimota -o out.xlsx
    python3 koscom_model_catalog.py --ajx <token>          # debug: skip auto-extract
"""
import argparse
import re
from collections import OrderedDict

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from koscom_common import KOSCOM_BASE_URL, make_session

MAKES = ["Honda", "Suzuki", "Kawasaki", "Yamaha", "Bimota", "BMW", "Aprilia"]
STATS_URL = f"{KOSCOM_BASE_URL}/bike_st"
HTTP_TIMEOUT = 20

# Candidate shapes for the ajx token in the stats-page source, tried in order.
# UNVERIFIED — extend once the real page can be viewed. Token chars: koscom
# tokens are typically [A-Za-z0-9] with maybe . _ -; keep the class tight so we
# don't grab surrounding markup.
_TOK = r"([A-Za-z0-9._-]{6,})"
AJX_TOKEN_PATTERNS = [
    rf"""ajx['"]?\s*[:=]\s*['"]{_TOK}['"]""",     # ajx:'x' / ajx = "x" / var ajx='x'
    rf"""[?&]ajx={_TOK}""",                         # ...&ajx=x inside a URL string
    rf"""name=['"]ajx['"][^>]*value=['"]{_TOK}['"]""",   # <input name=ajx value=x>
    rf"""value=['"]{_TOK}['"][^>]*name=['"]ajx['"]""",   # value first, then name
]


def extract_ajx_token(html):
    """Return the current ajx token from the stats-page source, or None."""
    for pat in AJX_TOKEN_PATTERNS:
        m = re.search(pat, html)
        if m:
            return m.group(1)
    return None


def fetch_ajx_token(session):
    """Load the stats page fresh and extract its current ajx token (raises if
    not found — never falls back to a stale/hardcoded value)."""
    resp = session.get(STATS_URL, timeout=HTTP_TIMEOUT)
    resp.encoding = "utf-8"
    token = extract_ajx_token(resp.text)
    if not token:
        raise RuntimeError(
            "could not extract an ajx token from the bike_st page source — the "
            "token pattern likely changed; view-source the page and update "
            "AJX_TOKEN_PATTERNS (or pass --ajx for a one-off debug run).")
    return token


# Match every ModelName#Count pair anywhere in the callback payload. The name
# excludes the structural chars (# = quotes parens) so a wrapping callback like
# cb('A#1=B#2') can't bleed into the first name; count is the trailing digits.
_PAIR_RE = re.compile(r"([^#=\"'()]+?)#(\d+)")


def parse_model_counts(text):
    """[(model, count)] from a 'Name#Count=Name#Count=...' payload, deduped
    (summing counts) while preserving first-seen order."""
    out = OrderedDict()
    for name, count in _PAIR_RE.findall(text or ""):
        name = name.strip()
        if not name:
            continue
        out[name] = out.get(name, 0) + int(count)
    return list(out.items())


def fetch_models_for_make(session, make, token, house=None):
    """POST the catalog request for one make; return [(model, count)]."""
    url = f"{STATS_URL}?file=ajxModel&ajx={token}"
    data = {"manuf": make}
    if house:
        data["house"] = house       # UNVERIFIED param name/value — see module docstring
    headers = {"X-Requested-With": "XMLHttpRequest", "Referer": STATS_URL + "/"}
    resp = session.post(url, data=data, headers=headers, timeout=HTTP_TIMEOUT)
    resp.encoding = "utf-8"
    return parse_model_counts(resp.text)


# ---------------------------------------------------------------- workbook out
_ARIAL = "Arial"
_HEADER_FONT = Font(name=_ARIAL, bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill("solid", fgColor="404040")
_BODY_FONT = Font(name=_ARIAL)


def _style_header(ws, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(ncols)}1"


def build_workbook(results, makes, path, house=None):
    """results: {make: [(model, count), ...]}. Writes `path`."""
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = "Models"
    scope = f"BDS only (house={house})" if house else "full catalog (all houses)"
    ws.append(["Make", "Model (koscom naming)", "Listings"])
    for make in makes:
        for model, count in sorted(results.get(make, []), key=lambda mc: mc[0].lower()):
            ws.append([make, model, count])
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = _BODY_FONT
    _style_header(ws, 3)
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 12

    sm = wb.create_sheet("Summary")
    sm.append(["Make", "Distinct models", "Total listings"])
    for make in makes:
        pairs = results.get(make, [])
        sm.append([make, len(pairs), sum(c for _, c in pairs)])
    sm.append([])
    sm.append([f"Source: koscom bike_st ajxModel endpoint — {scope}"])
    for row in sm.iter_rows(min_row=2, max_row=1 + len(makes)):
        for cell in row:
            cell.font = _BODY_FONT
    _style_header(sm, 3)
    for col, w in zip("ABC", (12, 16, 16)):
        sm.column_dimensions[col].width = w

    wb.save(path)
    return path


def main():
    ap = argparse.ArgumentParser(description="Pull koscom's per-make model catalog (research).")
    ap.add_argument("--makes", nargs="+", default=MAKES, help="makes to query")
    ap.add_argument("-o", "--out", default="koscom_model_catalog.xlsx", help="output .xlsx path")
    ap.add_argument("--house", default=None,
                    help="BDS pre-filter value sent as house=<VALUE> (UNVERIFIED; e.g. BDS)")
    ap.add_argument("--ajx", default=None,
                    help="debug only: use this ajx token instead of auto-extracting "
                         "(tokens expire — normal runs read it fresh)")
    args = ap.parse_args()

    session = make_session()
    token = args.ajx or fetch_ajx_token(session)
    if args.ajx:
        print("[WARN] using --ajx override; tokens are page-load specific and expire")
    print(f"[TOKEN] using ajx={token}")

    results = {}
    for make in args.makes:
        try:
            pairs = fetch_models_for_make(session, make, token, house=args.house)
        except Exception as e:
            print(f"[ERROR] {make}: {e}")
            pairs = []
        results[make] = pairs
        note = "" if pairs else "  (0 — token expired? param wrong? make unknown?)"
        print(f"[{make}] {len(pairs)} distinct models{note}")

    path = build_workbook(results, args.makes, args.out, house=args.house)
    total = sum(len(v) for v in results.values())
    print(f"\n[SAVED] {path} — {total} distinct model(s) across {len(args.makes)} make(s)")


if __name__ == "__main__":
    main()
