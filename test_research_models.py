#!/usr/bin/env python3
"""
Self-tests for research_koscom_models.py — no live access. Mocks koscom to
verify: browse pagination is consumed, BDS-only filtering, model extraction +
dedupe with counts, the "(model not found)" fallback, and that a valid .xlsx is
written with the expected sheets/rows.

    python3 test_research_models.py
"""
import os
import tempfile
import types
from urllib.parse import urlparse, parse_qs

import openpyxl

import koscom_scraper_v3 as sc
import research_koscom_models as rm

sc.REQUEST_DELAY_SECONDS = 0


def _search_html(ids):
    return "<html><body>" + "".join(
        f'<a href="/bike-{i}.htm">t</a><a href="/bike-{i}.htm">p</a>' for i in ids
    ) + "</body></html>"


def _detail(model=None, bds=True):
    house = "BDS Kantou" if bds else "Yahoo Auc"
    row = f'<tr><td class="bkth">車名</td><td>{model}</td></tr>' if model else \
          '<tr><td class="bkth">grade</td><td>4</td></tr>'
    return f"<html><body>{house} <table>{row}</table></body></html>"

# A=CBR400RR(BDS), B=CBR400RR(BDS dup), C=non-BDS, D=VFR400R(BDS),
# E=BDS but no model row -> "(model not found)"
DETAILS = {
    "A": _detail("CBR400RR"), "B": _detail("CBR400RR"),
    "C": _detail("CBR400RR", bds=False),
    "D": _detail("VFR400R"), "E": _detail(None),
}


def _make_scraper():
    def get(url, timeout=None):
        if "manuf=" in url:
            page = int(parse_qs(urlparse(url).query)["page"][0])
            ids = list(DETAILS) if page == 1 else []
            return types.SimpleNamespace(text=_search_html(ids), encoding="utf-8")
        lid = url.split("bike-")[1].split(".htm")[0]
        return types.SimpleNamespace(text=DETAILS[lid], encoding="utf-8")
    s = KoscomScraperV3 = sc.KoscomScraperV3()
    s.session = types.SimpleNamespace(get=get)
    s.bikes = []
    s.seen_listing_ids = set()
    return s


def test_collect_dedupes_and_counts():
    s = _make_scraper()
    r = rm.collect_models_for_make(s, "Honda", fetch_delay=0, log=lambda *a: None)
    assert r["models"]["CBR400RR"] == 2, r["models"]
    assert r["models"]["VFR400R"] == 1
    assert r["models"][rm.MODEL_NOT_FOUND] == 1
    assert r["scanned"] == 4          # A, B, D, E (C is non-BDS)
    assert r["non_bds"] == 1          # C


def test_browse_query_uses_force_flag():
    seen = []
    s = _make_scraper()
    base = s.session.get
    s.session = types.SimpleNamespace(get=lambda u, timeout=None: (seen.append(u), base(u, timeout))[1])
    rm.collect_models_for_make(s, "Bimota", fetch_delay=0, log=lambda *a: None)
    search_urls = [u for u in seen if "manuf=" in u]
    assert search_urls and all("manuf=Bimota" in u and "force=1" in u for u in search_urls), search_urls
    assert all("max_year=" not in u and "model=" not in u for u in search_urls), search_urls


def test_stale_limit_stops_early():
    s = _make_scraper()
    # stale_limit=1: after the first duplicate (B repeats CBR400RR) it stops.
    r = rm.collect_models_for_make(s, "Honda", stale_limit=1, fetch_delay=0, log=lambda *a: None)
    assert r["stopped_early"] is True
    assert r["scanned"] < 4           # did not scan the whole make


def test_build_workbook_structure():
    results = {
        "Honda": {"models": {"CBR400RR": 2, "VFR400R": 1, rm.MODEL_NOT_FOUND: 1},
                  "scanned": 4, "non_bds": 1, "stopped_early": False},
        "Bimota": {"models": {}, "scanned": 0, "non_bds": 0, "stopped_early": False},
    }
    with tempfile.TemporaryDirectory() as d:
        path = os.path.join(d, "out.xlsx")
        rm.build_workbook(results, ["Honda", "Bimota"], path)
        wb = openpyxl.load_workbook(path)
        assert wb.sheetnames == ["Models", "Summary"], wb.sheetnames
        ws = wb["Models"]
        assert [c.value for c in ws[1]] == ["Make", "Model (koscom naming)", "Listings seen"]
        rows = [(r[0].value, r[1].value, r[2].value) for r in ws.iter_rows(min_row=2)]
        # Honda models sorted alpha, not-found last; Bimota has no rows.
        assert rows == [
            ("Honda", "CBR400RR", 2),
            ("Honda", "VFR400R", 1),
            ("Honda", rm.MODEL_NOT_FOUND, 1),
        ], rows
        assert ws["A2"].font.name == "Arial"
        sm = wb["Summary"]
        smrows = {r[0].value: r[1].value for r in sm.iter_rows(min_row=2)}
        assert smrows["Honda"] == 3 and smrows["Bimota"] == 0


if __name__ == "__main__":
    fns = [v for k, v in sorted(globals().items()) if k.startswith("test_")]
    for fn in fns:
        fn()
        print(f"ok  {fn.__name__}")
    print(f"\n{len(fns)} test(s) passed")
